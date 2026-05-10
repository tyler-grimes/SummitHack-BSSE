"""
Seed historical ERCOT RT LMP from the public MIS annual archive (no API key required).

Data source:
  ERCOT MIS — Report 13061 (Historical RTM Load Zone and Hub Prices, 15-min SPP)
  One ZIP per year, each containing an XLSX with 12 monthly sheets.
  Columns: Delivery Date, Delivery Hour, Delivery Interval, Repeated Hour Flag,
           Settlement Point Name, Settlement Point Type, Settlement Point Price

  Doc IDs (stable, verified 2026-05-10):
    2020 → 751356439
    2021 → 814922832
    2022 → 886632075
    2023 → 969805139
    2024 → 1065471230
    (2025 and 2026 are already covered by GridStatus seed)

Scope:
  Downloads 5 annual files (~14 MB each) covering 2020–2024.
  Filters to HB_NORTH, HB_SOUTH, HB_WEST, HB_HOUSTON only.
  Inserts into the lmp table. ON CONFLICT DO NOTHING — idempotent.

Usage:
  cd /mnt/c/Users/tygri/energy-trading-optimization
  packages/data-pipeline/.venv/bin/python scripts/seed_mis_historical.py

  To seed a single year:
  YEARS=2022 packages/data-pipeline/.venv/bin/python scripts/seed_mis_historical.py
"""

import asyncio
import io
import logging
import os
import sys
import zipfile
from datetime import datetime, timezone

import httpx
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "packages", "data-pipeline"))

from dotenv import load_dotenv
load_dotenv(os.path.join(ROOT, ".env"))

from src.db import init_pool, close_pool, insert_lmp_batch

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

HUBS = {"HB_NORTH", "HB_SOUTH", "HB_WEST", "HB_HOUSTON"}

# Stable MIS doc IDs for annual RT SPP files (report 13061).
# These are permanent archive IDs — they don't change.
ANNUAL_DOC_IDS: dict[int, int] = {
    2020: 751356439,
    2021: 814922832,
    2022: 886632075,
    2023: 969805139,
    2024: 1065471230,
}

MIS_DOWNLOAD_URL = (
    "https://www.ercot.com/misdownload/servlets/mirDownload"
    "?mimic_duns=000000000&doclookupId={doc_id}"
)

HEADERS = {"User-Agent": "energy-trading-bess-seed/1.0 (research)"}

# Batch size for DB inserts — keeps memory bounded on large sheets
INSERT_BATCH = 5_000

# ---------------------------------------------------------------------------
# XLSX parsing
# ---------------------------------------------------------------------------

# ERCOT uses Central time. UTC offsets:
#   CDT (Apr–Oct): UTC-5   CST (Nov–Mar): UTC-6
# We apply a simple month-based heuristic — good enough for hub prices.
def _utc_offset(month: int) -> int:
    return -5 if 4 <= month <= 10 else -6


def _parse_xlsx(xlsx_bytes: bytes, year: int) -> list[dict]:
    """
    Parse an annual ERCOT RTM SPP xlsx and return insert-ready dicts.
    Each row: Delivery Date, Hour (1-24), Interval (1-4), Repeated Hour Flag,
              Settlement Point Name, Settlement Point Type, Price
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    records: list[dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)

        # Skip header row
        try:
            header = next(rows)
        except StopIteration:
            continue

        # Normalise header to find column positions robustly
        h = [str(c).strip().lower() if c else "" for c in header]
        def idx(fragment: str) -> int | None:
            for i, col in enumerate(h):
                if fragment in col:
                    return i
            return None

        date_i  = idx("delivery date")
        hour_i  = idx("delivery hour")
        intv_i  = idx("delivery interval")
        name_i  = idx("settlement point name")
        price_i = idx("settlement point price")

        if any(i is None for i in [date_i, hour_i, intv_i, name_i, price_i]):
            log.warning("Sheet '%s': unexpected header %s — skipping", sheet_name, header)
            continue

        sheet_rows = 0
        for row in rows:
            if row[name_i] is None:
                continue
            node = str(row[name_i]).strip().upper()
            if node not in HUBS:
                continue

            try:
                raw_date = row[date_i]      # MM/DD/YYYY string or date object
                hour     = int(row[hour_i])  # 1-24
                interval = int(row[intv_i])  # 1-4
                price    = float(row[price_i])
            except (TypeError, ValueError):
                continue

            # Parse date
            if isinstance(raw_date, str):
                parts = raw_date.strip().split("/")
                if len(parts) == 3:
                    month, day, yr = int(parts[0]), int(parts[1]), int(parts[2])
                else:
                    continue
            elif hasattr(raw_date, "month"):
                month, day, yr = raw_date.month, raw_date.day, raw_date.year
            else:
                continue

            # Convert ERCOT hour (1-24) + interval (1-4) → UTC timestamp
            hour_0    = hour - 1           # 0-based
            minute    = (interval - 1) * 15
            utc_off   = _utc_offset(month)

            try:
                # Build a timezone-aware datetime in Central time then convert to UTC
                local_naive = datetime(yr, month, day, hour_0, minute, 0)
                utc_dt = local_naive.replace(
                    tzinfo=timezone.utc
                ) - __import__("datetime").timedelta(hours=utc_off)
                # Simpler: store with explicit offset string for asyncpg
                ts = f"{yr:04d}-{month:02d}-{day:02d}T{hour_0:02d}:{minute:02d}:00{utc_off:+03d}:00"
            except ValueError:
                continue  # e.g. hour=24 DST transition edge case

            records.append({
                "timestamp": ts,
                "iso": "ERCOT",
                "node": node,
                "lmp": price,
                "energy": price,
                "congestion": 0.0,
                "loss": 0.0,
            })
            sheet_rows += 1

        log.info("  Sheet %-3s: %d hub records parsed", sheet_name, sheet_rows)

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

async def seed_year(client: httpx.AsyncClient, year: int, doc_id: int) -> int:
    """Download, parse, and insert one annual file. Returns rows inserted."""
    url = MIS_DOWNLOAD_URL.format(doc_id=doc_id)
    log.info("=== Year %d  downloading from MIS (docId=%d) ===", year, doc_id)

    resp = await client.get(url, timeout=120)
    resp.raise_for_status()
    log.info("  Downloaded %.1f MB", len(resp.content) / 1_048_576)

    # Extract XLSX from ZIP
    with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            log.error("  No XLSX found in ZIP for year %d", year)
            return 0
        xlsx_bytes = zf.read(xlsx_names[0])
        log.info("  Extracted %s (%.1f MB)", xlsx_names[0], len(xlsx_bytes) / 1_048_576)

    log.info("  Parsing XLSX...")
    records = _parse_xlsx(xlsx_bytes, year)
    log.info("  Parsed %d hub records for year %d", len(records), year)

    if not records:
        return 0

    # Insert in batches to keep memory bounded
    total = 0
    for i in range(0, len(records), INSERT_BATCH):
        batch = records[i : i + INSERT_BATCH]
        inserted = await insert_lmp_batch(batch)
        total += inserted

    log.info("=== Year %d done: %d rows inserted ===", year, total)
    return total


async def main() -> None:
    # Allow YEARS=2022,2023 env var to seed a subset
    years_env = os.environ.get("YEARS")
    if years_env:
        target_years = [int(y.strip()) for y in years_env.split(",")]
    else:
        target_years = sorted(ANNUAL_DOC_IDS.keys())

    log.info("Connecting to DB...")
    await init_pool()
    log.info("Seeding years: %s", target_years)

    grand_total = 0
    async with httpx.AsyncClient(headers=HEADERS, follow_redirects=True) as client:
        for year in target_years:
            doc_id = ANNUAL_DOC_IDS.get(year)
            if doc_id is None:
                log.warning("No doc ID known for year %d — skipping", year)
                continue
            try:
                n = await seed_year(client, year, doc_id)
                grand_total += n
            except Exception as exc:
                log.error("Year %d failed: %s", year, exc)

    await close_pool()
    log.info("All done. Total rows inserted: %d", grand_total)


if __name__ == "__main__":
    asyncio.run(main())
